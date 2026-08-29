from pypdf import PdfReader
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule

def parse_amount(amount_text):
    if amount_text is None:
        return None
    cleaned = amount_text.replace(" ", "")
    cleaned = cleaned.replace(",", ".")
    return float(cleaned)

def extract_invoice_data(pdf_path):
    reader = PdfReader(pdf_path)

    page = reader.pages[0]
    text = page.extract_text()
    lines = text.splitlines()

    invoice_number = None
    invoice_date = None
    amount_due = None
    net_amount = None
    low_tax_value = None
    high_tax_value = None
    gross_amount = None

    for index, line in enumerate(lines):
        if line == "Számla száma":
            invoice_number = lines[index + 1]
        
        if line == "Számla kelte":
            invoice_date = lines[index+1]

        if line.startswith("Számlaérték adó nélkül:"):
            net_amount = line.split(":", 1)[1].strip()

        if line.startswith("5 %-os adóalap"):
            low_tax_value = line.split("Adóérték:", 1)[1].strip()

        if line.startswith("27 %-os adóalap"):
            high_tax_value = line.split("Adóérték:", 1)[1].strip()

        if line.startswith("Számlaérték összesen"):
            gross_amount = line.split("összesen:", 1)[1].strip()

        if line.startswith("Fizetendő:"):
            amount_due = line.split(":", 1)[1].strip()

    net_amount = parse_amount(net_amount)
    gross_amount = parse_amount(gross_amount)
    amount_due = parse_amount(amount_due)
    low_tax_value = parse_amount(low_tax_value)
    high_tax_value = parse_amount(high_tax_value)

    if low_tax_value is None:
        low_tax_value = 0

    if high_tax_value is None:
        high_tax_value = 0

    total_vat = low_tax_value + high_tax_value

    if net_amount is not None and gross_amount is not None:
        difference = abs((net_amount + total_vat) - gross_amount)
        accounting_check = difference < 0.01
    else:
        accounting_check = False

    return {
        "source_file": pdf_path.name,
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "net_amount": net_amount,
        "vat_5": low_tax_value,
        "vat_27": high_tax_value,
        "total_vat": total_vat,
        "gross_amount": gross_amount,
        "amount_due": amount_due,
        "accounting_check": accounting_check
        }

invoice_folder = Path("invoices")
all_invoices = []

for file in invoice_folder.iterdir():
    if file.suffix.lower() != ".pdf":
        continue

    invoice_data = extract_invoice_data(file)
    all_invoices.append(invoice_data)

for invoice in all_invoices:
    print(invoice)

workbook = Workbook()
sheet = workbook.active
sheet.title = "Invoice Summary"

sheet.append([
    "Source File",
    "Invoice Number",
    "Invoice Date",
    "Net Amount",
    "VAT 5%",
    "VAT 27%",
    "Total VAT",
    "Gross Amount",
    "Amount Due",
    "Accounting Check"
])

for invoice in all_invoices:
    sheet.append([
        invoice["source_file"],
        invoice["invoice_number"],
        invoice["invoice_date"],
        invoice["net_amount"],
        invoice["vat_5"],
        invoice["vat_27"],
        invoice["total_vat"],
        invoice["gross_amount"],
        invoice["amount_due"],
        invoice["accounting_check"]
    ])

huf_format = '#,##0.00 "HUF"'

thin = Side(style="thin")
border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

for cell in sheet[1]:
    cell.font = Font(bold=True)
    cell.border = border
    cell.alignment = Alignment(horizontal="center")

for row in range(2, sheet.max_row + 1):
    sheet[f"D{row}"].number_format = huf_format
    sheet[f"E{row}"].number_format = huf_format
    sheet[f"F{row}"].number_format = huf_format
    sheet[f"G{row}"].number_format = huf_format
    sheet[f"H{row}"].number_format = huf_format
    sheet[f"I{row}"].number_format = huf_format

for row in sheet.iter_rows():
    for cell in row:
        cell.border = border

sheet.column_dimensions["A"].width = 28
sheet.column_dimensions["B"].width = 20
sheet.column_dimensions["C"].width = 16
sheet.column_dimensions["D"].width = 18
sheet.column_dimensions["E"].width = 14
sheet.column_dimensions["F"].width = 14
sheet.column_dimensions["G"].width = 16
sheet.column_dimensions["H"].width = 18
sheet.column_dimensions["I"].width = 18
sheet.column_dimensions["J"].width = 18

sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions

for row in range(2, sheet.max_row + 1):
    sheet[f"J{row}"].alignment = Alignment(horizontal="center")

green_fill = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

red_fill = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)

check_range = f"J2:J{sheet.max_row}"

sheet.conditional_formatting.add(
    check_range,
    CellIsRule(
        operator="equal",
        formula=["TRUE"],
        fill=green_fill
    )
)

sheet.conditional_formatting.add(
    check_range,
    CellIsRule(
        operator="equal",
        formula=["FALSE"],
        fill=red_fill
    )
)

workbook.save("invoice_summary.xlsx")